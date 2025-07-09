#!/usr/bin/env python3
"""
Create a simple test Excel file using openpyxl (if available) or xlsxwriter
"""
import sys

def create_with_openpyxl():
    """Create test Excel file using openpyxl"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add header row
        ws.append(['Name', 'Age', 'City', 'Salary'])
        
        # Add data rows
        ws.append(['John Doe', 25, 'New York', 50000])
        ws.append(['Jane Smith', 30, 'Los Angeles', 60000])
        ws.append(['Bob Johnson', 35, 'Chicago', 70000])
        
        # Create a second sheet
        ws2 = wb.create_sheet("Products")
        ws2.append(['Product', 'Price', 'Stock'])
        ws2.append(['Laptop', 999.99, 10])
        ws2.append(['Mouse', 25.50, 50])
        ws2.append(['Keyboard', 75.00, 25])
        
        wb.save('test_data.xlsx')
        print("✅ Created test_data.xlsx using openpyxl")
        return True
    except ImportError:
        return False

def create_with_xlsxwriter():
    """Create test Excel file using xlsxwriter"""
    try:
        import xlsxwriter
        
        workbook = xlsxwriter.Workbook('test_data.xlsx')
        
        # First sheet
        worksheet1 = workbook.add_worksheet('Sheet1')
        worksheet1.write_row(0, 0, ['Name', 'Age', 'City', 'Salary'])
        worksheet1.write_row(1, 0, ['John Doe', 25, 'New York', 50000])
        worksheet1.write_row(2, 0, ['Jane Smith', 30, 'Los Angeles', 60000])
        worksheet1.write_row(3, 0, ['Bob Johnson', 35, 'Chicago', 70000])
        
        # Second sheet
        worksheet2 = workbook.add_worksheet('Products')
        worksheet2.write_row(0, 0, ['Product', 'Price', 'Stock'])
        worksheet2.write_row(1, 0, ['Laptop', 999.99, 10])
        worksheet2.write_row(2, 0, ['Mouse', 25.50, 50])
        worksheet2.write_row(3, 0, ['Keyboard', 75.00, 25])
        
        workbook.close()
        print("✅ Created test_data.xlsx using xlsxwriter")
        return True
    except ImportError:
        return False

def create_csv_fallback():
    """Create CSV files as fallback"""
    with open('test_data.csv', 'w') as f:
        f.write('Name,Age,City,Salary\n')
        f.write('John Doe,25,New York,50000\n')
        f.write('Jane Smith,30,Los Angeles,60000\n')
        f.write('Bob Johnson,35,Chicago,70000\n')
    
    print("✅ Created test_data.csv as fallback")
    print("Note: You'll need a real Excel file to test the xlsx parser")

def main():
    print("Creating test Excel file...")
    
    if create_with_openpyxl():
        return
    
    if create_with_xlsxwriter():
        return
    
    print("❌ Neither openpyxl nor xlsxwriter available")
    print("Install one of them: pip install openpyxl  or  pip install xlsxwriter")
    
    create_csv_fallback()

if __name__ == "__main__":
    main()